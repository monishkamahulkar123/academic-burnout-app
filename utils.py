import streamlit as st
from datetime import datetime, timedelta
import plotly.graph_objects as go
import plotly.express as px
from fpdf import FPDF
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def apply_custom_css():
    """Apply minimal professional styling"""
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    
    * {
        font-family: 'Inter', sans-serif;
    }
    
    .stApp {
        background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
    }
    
    h1 {
        color: #1e3a8a;
        font-weight: 700;
        padding: 20px 0;
        border-bottom: 4px solid #3b82f6;
        margin-bottom: 30px;
    }
    
    h2 {
        color: #1e40af;
        font-weight: 600;
        margin-top: 20px;
    }
    
    h3 {
        color: #2563eb;
        font-weight: 500;
    }
    
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1e3a8a 0%, #3b82f6 100%);
    }
    
    [data-testid="stSidebar"] .stMarkdown {
        color: white !important;
    }
    
    [data-testid="stSidebar"] button {
        background-color: rgba(255,255,255,0.1);
        color: white;
        border: 1px solid rgba(255,255,255,0.3);
        transition: all 0.3s;
    }
    
    [data-testid="stSidebar"] button:hover {
        background-color: rgba(255,255,255,0.2);
        border-color: white;
        transform: translateX(5px);
    }
    
    .stButton>button {
        background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 10px 24px;
        font-weight: 600;
        transition: all 0.3s;
        box-shadow: 0 4px 6px rgba(59, 130, 246, 0.3);
    }
    
    .stButton>button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 12px rgba(59, 130, 246, 0.4);
    }
    
    [data-testid="stMetricValue"] {
        font-size: 32px;
        font-weight: 700;
        color: #1e3a8a;
    }
    
    [data-testid="stMetricLabel"] {
        font-size: 14px;
        font-weight: 500;
        color: #64748b;
    }
    
    .streamlit-expanderHeader {
        background-color: white;
        border-radius: 12px;
        border: 2px solid #e2e8f0;
        font-weight: 600;
        transition: all 0.3s;
    }
    
    .streamlit-expanderHeader:hover {
        border-color: #3b82f6;
        box-shadow: 0 4px 12px rgba(59, 130, 246, 0.2);
    }
    
    .stForm {
        background-color: white;
        padding: 30px;
        border-radius: 15px;
        box-shadow: 0 8px 16px rgba(0,0,0,0.1);
        border: 1px solid #e2e8f0;
    }
    
    .stSuccess {
        background-color: #d1fae5;
        color: #065f46;
        border-radius: 10px;
        border-left: 4px solid #10b981;
        padding: 15px;
    }
    
    .stError {
        background-color: #fee2e2;
        color: #991b1b;
        border-radius: 10px;
        border-left: 4px solid #ef4444;
        padding: 15px;
    }
    
    .stWarning {
        background-color: #fef3c7;
        color: #92400e;
        border-radius: 10px;
        border-left: 4px solid #f59e0b;
        padding: 15px;
    }
    
    .stInfo {
        background-color: #dbeafe;
        color: #1e40af;
        border-radius: 10px;
        border-left: 4px solid #3b82f6;
        padding: 15px;
    }
    
    div[data-testid="stExpander"] {
        background-color: white;
        border-radius: 12px;
        margin-bottom: 15px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }
    
    .task-card {
        background: white;
        padding: 20px;
        border-radius: 12px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.07);
        margin: 10px 0;
        border-left: 4px solid #3b82f6;
        transition: all 0.3s;
    }
    
    .task-card:hover {
        box-shadow: 0 8px 16px rgba(0,0,0,0.12);
        transform: translateY(-2px);
    }
    
    [data-testid="stProgress"] > div > div > div {
        background: linear-gradient(90deg, #3b82f6 0%, #8b5cf6 100%);
    }
    </style>
    """, unsafe_allow_html=True)

def show_logo():
    """Display app logo and title"""
    st.markdown("""
    <div style='text-align: center; padding: 20px; background: white; border-radius: 15px; margin-bottom: 30px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);'>
        <h1 style='margin: 0; color: #1e3a8a; border: none;'>
            📚 Academic Burnout Detector
        </h1>
        <p style='color: #64748b; margin-top: 10px; font-size: 16px;'>
            Smart Task Management & Workload Analysis
        </p>
    </div>
    """, unsafe_allow_html=True)

def create_progress_bar(completed, total):
    """Create a visual progress bar"""
    if total == 0:
        percentage = 0
    else:
        percentage = int((completed / total) * 100)
    
    if percentage >= 70:
        color = '#10b981'
    elif percentage >= 40:
        color = '#f59e0b'
    else:
        color = '#ef4444'
    
    html = f"""
    <div style="background-color: #e5e7eb; border-radius: 12px; height: 30px; width: 100%; overflow: hidden; box-shadow: inset 0 2px 4px rgba(0,0,0,0.1);">
        <div style="background: linear-gradient(90deg, {color} 0%, {color}dd 100%); width: {percentage}%; height: 100%; 
                    display: flex; align-items: center; justify-content: center; color: white; font-weight: 700; font-size: 14px;
                    transition: width 0.5s ease;">
            {percentage}%
        </div>
    </div>
    """
    return html

def create_task_completion_chart(tasks):
    """Create pie chart for task completion"""
    if not tasks:
        return None
    
    completed = len([t for t in tasks if t.get('task_status') == 'Completed'])
    in_progress = len([t for t in tasks if t.get('task_status') == 'In Progress'])
    pending = len([t for t in tasks if t.get('task_status', 'Pending') == 'Pending'])
    
    fig = go.Figure(data=[go.Pie(
        labels=['Completed', 'In Progress', 'Pending'],
        values=[completed, in_progress, pending],
        hole=0.4,
        marker=dict(colors=['#10b981', '#f59e0b', '#ef4444']),
        textinfo='label+percent',
        textfont=dict(size=14, color='white', family='Inter')
    )])
    
    fig.update_layout(
        title=dict(text='Task Completion Status', font=dict(size=18, color='#1e3a8a', family='Inter')),
        showlegend=True,
        height=400,
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)'
    )
    
    return fig

def create_priority_distribution_chart(tasks):
    """Create bar chart for priority distribution"""
    if not tasks:
        return None
    
    priorities = {'Low': 0, 'Medium': 0, 'High': 0}
    for task in tasks:
        priority = task.get('priority', 'Low')
        priorities[priority] = priorities.get(priority, 0) + 1
    
    fig = go.Figure(data=[go.Bar(
        x=list(priorities.keys()),
        y=list(priorities.values()),
        marker=dict(color=['#10b981', '#f59e0b', '#ef4444']),
        text=list(priorities.values()),
        textposition='auto',
        textfont=dict(size=16, color='white', family='Inter')
    )])
    
    fig.update_layout(
        title=dict(text='Task Priority Distribution', font=dict(size=18, color='#1e3a8a', family='Inter')),
        xaxis=dict(title='Priority Level', title_font=dict(size=14, color='#64748b')),
        yaxis=dict(title='Number of Tasks', title_font=dict(size=14, color='#64748b')),
        height=400,
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)'
    )
    
    return fig

def create_workload_timeline(tasks):
    """Create timeline chart showing workload over time"""
    if not tasks:
        return None
    
    task_dates = {}
    for task in tasks:
        date = task['deadline']
        if date not in task_dates:
            task_dates[date] = 0
        task_dates[date] += task['estimated_hours']
    
    sorted_dates = sorted(task_dates.keys())
    hours = [task_dates[date] for date in sorted_dates]
    
    fig = go.Figure(data=[go.Scatter(
        x=sorted_dates,
        y=hours,
        mode='lines+markers',
        line=dict(color='#3b82f6', width=3),
        marker=dict(size=10, color='#1e3a8a'),
        fill='tozeroy',
        fillcolor='rgba(59, 130, 246, 0.2)'
    )])
    
    fig.update_layout(
        title=dict(text='Workload Timeline', font=dict(size=18, color='#1e3a8a', family='Inter')),
        xaxis=dict(title='Deadline Date', title_font=dict(size=14, color='#64748b')),
        yaxis=dict(title='Total Hours', title_font=dict(size=14, color='#64748b')),
        height=400,
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)'
    )
    
    return fig

def export_tasks_to_excel(tasks, filename="tasks_export.xlsx"):
    """Export tasks to Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    
    # Header styling
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    headers = ["Task Title", "Deadline", "Estimated Hours", "Priority", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for row, task in enumerate(tasks, 2):
        ws.cell(row=row, column=1, value=task.get('title', ''))
        ws.cell(row=row, column=2, value=str(task.get('deadline', '')))
        ws.cell(row=row, column=3, value=task.get('estimated_hours', 0))
        ws.cell(row=row, column=4, value=task.get('priority', ''))
        ws.cell(row=row, column=5, value=task.get('task_status', 'Pending'))
    
    # Adjust column widths
    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def export_report_to_pdf(username, tasks, group_tasks, burnout_data):
    """Generate PDF report"""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 20)
    
    # Title
    pdf.cell(0, 15, "Academic Workload Report", ln=True, align="C")
    pdf.ln(5)
    
    # User info
    pdf.set_font("Arial", "", 12)
    pdf.cell(0, 10, f"Student: {username}", ln=True)
    pdf.cell(0, 10, f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}", ln=True)
    pdf.ln(5)
    
    # Summary
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, "Summary", ln=True)
    pdf.set_font("Arial", "", 11)
    pdf.cell(0, 8, f"Total Individual Tasks: {len(tasks)}", ln=True)
    pdf.cell(0, 8, f"Total Group Tasks: {len(group_tasks)}", ln=True)
    
    risk_level, score, total, due_week, hours = burnout_data
    pdf.cell(0, 8, f"Burnout Risk: {risk_level} (Score: {score})", ln=True)
    pdf.cell(0, 8, f"Tasks Due This Week: {due_week}", ln=True)
    pdf.cell(0, 8, f"Estimated Hours This Week: {hours}", ln=True)
    pdf.ln(5)
    
    # Individual Tasks
    if tasks:
        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, "Individual Tasks", ln=True)
        pdf.set_font("Arial", "", 10)
        
        for task in tasks[:10]:  # Limit to 10 tasks
            pdf.cell(0, 7, f"- {task['title']} | Due: {task['deadline']} | {task['priority']} Priority", ln=True)
    
    # Output
    output = io.BytesIO()
    pdf_content = pdf.output(dest='S').encode('latin1')
    output.write(pdf_content)
    output.seek(0)
    
    return output

def show_notification(message, icon="🔔"):
    """Display notification banner"""
    st.markdown(f"""
    <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                color: white; padding: 15px 20px; border-radius: 10px; margin: 10px 0;
                box-shadow: 0 4px 6px rgba(102, 126, 234, 0.4); display: flex; align-items: center;'>
        <span style='font-size: 24px; margin-right: 15px;'>{icon}</span>
        <span style='font-size: 15px; font-weight: 500;'>{message}</span>
    </div>
    """, unsafe_allow_html=True)

def create_calendar_view(tasks, year, month):
    """Create monthly calendar view"""
    import calendar
    
    cal = calendar.monthcalendar(year, month)
    month_name = calendar.month_name[month]
    
    # Group tasks by date
    task_dict = {}
    for task in tasks:
        if task['deadline'].year == year and task['deadline'].month == month:
            day = task['deadline'].day
            if day not in task_dict:
                task_dict[day] = []
            task_dict[day].append(task)
    
    html = f"""
    <div style='background: white; padding: 20px; border-radius: 15px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);'>
        <h2 style='text-align: center; color: #1e3a8a; margin-bottom: 20px;'>{month_name} {year}</h2>
        <table style='width: 100%; border-collapse: collapse;'>
            <tr style='background: #3b82f6; color: white;'>
                <th style='padding: 10px; border: 1px solid #ddd;'>Mon</th>
                <th style='padding: 10px; border: 1px solid #ddd;'>Tue</th>
                <th style='padding: 10px; border: 1px solid #ddd;'>Wed</th>
                <th style='padding: 10px; border: 1px solid #ddd;'>Thu</th>
                <th style='padding: 10px; border: 1px solid #ddd;'>Fri</th>
                <th style='padding: 10px; border: 1px solid #ddd;'>Sat</th>
                <th style='padding: 10px; border: 1px solid #ddd;'>Sun</th>
            </tr>
    """
    
    for week in cal:
        html += "<tr>"
        for day in week:
            if day == 0:
                html += "<td style='padding: 15px; border: 1px solid #ddd; background: #f9fafb;'></td>"
            else:
                task_count = len(task_dict.get(day, []))
                bg_color = '#fee2e2' if task_count > 2 else '#fef3c7' if task_count > 0 else 'white'
                html += f"<td style='padding: 15px; border: 1px solid #ddd; background: {bg_color}; vertical-align: top;'>"
                html += f"<strong>{day}</strong>"
                if task_count > 0:
                    html += f"<br><span style='color: #ef4444; font-size: 12px;'>📋 {task_count} task(s)</span>"
                html += "</td>"
        html += "</tr>"
    
    html += "</table></div>"
    
    return html